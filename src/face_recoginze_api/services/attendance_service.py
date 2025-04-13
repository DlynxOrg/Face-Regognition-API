from sqlmodel.ext.asyncio.session import AsyncSession
from face_recoginze_api.DTOs.dtos import AttendanceRecordCreate, AttendanceRecordRead
from ..repositories.attendance_repository import AttendanceRepository
from typing import List
from datetime import timedelta
from ..DTOs.dtos import UserDTO, CreditClassRead
import pandas as pd
from pathlib import Path
from datetime import datetime
from ..enums.enums import EXCEL_FILE_PATH
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
class AttendanceService:
    def __init__(self, repo: AttendanceRepository):
        self.repo = repo

    async def create_attendance(self, db: AsyncSession, data: AttendanceRecordCreate) -> AttendanceRecordRead:
        # Tạo bản ghi điểm danh mới
        record = await self.repo.create(db, data)

        # Trả về bản ghi điểm danh đã tạo dưới dạng AttendanceRecordRead
        return AttendanceRecordRead.model_validate({
                **record.__dict__,
                "timestamp": record.timestamp + timedelta(hours=7),
            })

    async def get_all_attendance(self, db: AsyncSession) -> List[AttendanceRecordRead]:
        records = await self.repo.get_all(db)
        return [
            AttendanceRecordRead.model_validate({
                **record.__dict__,
                "timestamp": record.timestamp + timedelta(hours=7),
            })
            for record in records
        ]

    async def get_attendance_by_user_id(self, db: AsyncSession, user_id: int) -> List[AttendanceRecordRead]:
        records = await self.repo.get_by_user_id(db, user_id)
        return [
            AttendanceRecordRead.model_validate({
                **record.__dict__,
                "timestamp": record.timestamp + timedelta(hours=7),
            })
            for record in records
        ]

    async def get_attendance_by_class_id(self, db: AsyncSession, class_id: int) -> List[AttendanceRecordRead]:
        records = await self.repo.get_by_class_id(db, class_id)
        return [
            AttendanceRecordRead.model_validate({
                **record.__dict__,
                "timestamp": record.timestamp + timedelta(hours=7),
            })
            for record in records
        ]
    
    async def export_to_excel(self, records: list[AttendanceRecordRead]) -> Path:
        data = []
        for record in records:
            data.append({
                "Mã sinh viên": record.user.id,
                "Tên sinh viên": record.user.name,
                "Lớp học phần": record.credit_class.name,
                "Thời gian điểm danh": record.timestamp.strftime("%H:%M - %d/%m/%Y")
            })

        df = pd.DataFrame(data)
        filename = f"attendance_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = Path(EXCEL_FILE_PATH) / filename
        filepath.parent.mkdir(parents=True, exist_ok=True)  # Tạo folder nếu chưa có
        df.index += 1
        df.index.name = "STT"
        df.to_excel(filepath, index=True, engine="openpyxl")

        wb = load_workbook(filepath)
        ws = wb.active

        # Căn giữa toàn bộ ô
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # In đậm dòng tiêu đề và thêm viền quanh ô
        header_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for cell in ws[1]:
            cell.font = header_font
            cell.border = thin_border

        # Thêm viền cho các ô còn lại
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=len(df.columns) + 1, max_row=len(df) + 1):
            for cell in row:
                cell.border = thin_border

        # Tự động điều chỉnh chiều rộng cột
        for column_cells in ws.columns:
            max_length = 0
            col_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = adjusted_width

        # Freeze dòng đầu tiên (dòng tiêu đề)
        ws.freeze_panes = "A2"

        wb.save(filepath)
        return filepath

